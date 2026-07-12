# UmaCP - Umamusume Completionist Planner
# Copyright (C) 2026
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

import faulthandler
faulthandler.enable()

import streamlit as st
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from db import get_all_races
from scheduler import (
    optimize_schedule,
    get_build_requirements,
    RANK_SCALE,
    RANK_NAMES
)
from renderer import render_calendar_year

# Page configuration
st.set_page_config(
    page_title="Umamusume Completionist Planner",
    page_icon="🏇",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for premium dark-theme look & feel
st.html("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700;800&display=swap');

/* Global styles */
html, body, [class*="css"] {
    font-family: 'Outfit', sans-serif;
}

/* Header styling */
.main-header {
    background: linear-gradient(135deg, #1f1235 0%, #0f081d 100%);
    padding: 30px;
    border-radius: 15px;
    border: 1px solid rgba(255, 255, 255, 0.05);
    margin-bottom: 25px;
    box-shadow: 0 10px 30px rgba(0, 0, 0, 0.4);
    text-align: center;
}

.main-header h1 {
    margin: 0;
    font-size: 2.8rem;
    font-weight: 800;
    background: linear-gradient(90deg, #a855f7 0%, #ec4899 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    letter-spacing: -1px;
}

.main-header p {
    color: #a0aec0;
    margin-top: 10px;
    font-size: 1.1rem;
    font-weight: 300;
}

/* Custom styles for native elements */
div[data-testid="stContainer"] {
    background-color: rgba(255, 255, 255, 0.02);
    border: 1px solid rgba(255, 255, 255, 0.08);
    border-radius: 12px;
    padding: 24px;
    margin-bottom: 25px;
    box-shadow: 0 4px 20px rgba(0, 0, 0, 0.2);
    transition: all 0.3s ease;
}

div[data-testid="stContainer"]:hover {
    border-color: rgba(168, 85, 247, 0.4);
    box-shadow: 0 10px 30px rgba(168, 85, 247, 0.1);
}

/* Badges */
.stat-badge {
    background: rgba(168, 85, 247, 0.15);
    color: #c084fc;
    padding: 3px 8px;
    border-radius: 5px;
    border: 1px solid rgba(168, 85, 247, 0.3);
    font-size: 0.8rem;
    margin-right: 5px;
}

.summary-metric {
    font-size: 2.2rem;
    font-weight: 800;
    color: #a855f7;
}

</style>
""")



# Generate Excel spreadsheet
def generate_excel_file(careers, base_ranks):
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    for idx, career in enumerate(careers):
        sheet_name = f"Career {idx + 1}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Print build details at top
        build_reqs = get_build_requirements(career['build'])
        build_str = "No upgrades needed"
        if build_reqs['total_stars'] > 0:
            parts = []
            for attr, info in build_reqs['attributes'].items():
                base_val = base_ranks.get(attr, 0)
                final_val = min(6, base_val + info['steps'])
                final_letter = RANK_NAMES.get(final_val, 'G')
                parts.append(f"+{info['steps']} {attr.capitalize()} ({info['stars']}★, {info['slots']} slots, [{final_letter}])")
            build_str = ", ".join(parts)
            
        # Write metadata at the top
        ws['A1'] = "Inheritance Build:"
        ws['A1'].font = Font(bold=True, size=11, color="4B2375")
        ws['B1'] = build_str
        ws['B1'].font = Font(bold=True, size=11)
        
        ws['A2'] = "Total Stars Required:"
        ws['A2'].font = Font(color="555555")
        ws['B2'] = f"{build_reqs['total_stars']}★ (Max 18★)"
        
        ws['C2'] = "Total Slots Required:"
        ws['C2'].font = Font(color="555555")
        ws['D2'] = f"{build_reqs['total_slots']} slots (Max 6)"
        
        # Write table headers
        headers = ['Turn', 'Calendar', 'Race Name', 'Grade', 'Distance', 'Surface']
        header_fill = PatternFill(start_color="1F1235", end_color="1F1235", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            
        # Write schedule details row by row
        for turn in range(1, 73):
            race = career['schedule'][turn]
            
            year_name = "Junior" if turn <= 24 else ("Classic" if turn <= 48 else "Senior")
            turn_in_year = (turn - 1) % 24
            month = (turn_in_year // 2) + 1
            half = "Early" if (turn_in_year % 2) == 0 else "Late"
            
            month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            month_name = month_names[month - 1]
            time_str = f"{year_name} {month_name} {half}"
            
            if race:
                grade_str = ["G1", "G2", "G3"][race['grade']]
                dist_str = ["Sprint", "Mile", "Medium", "Long"][race['distance']]
                surf_str = ["Turf", "Dirt"][race['surface']]
                
                row_values = [turn, time_str, race['name'], grade_str, dist_str, surf_str]
            else:
                row_values = [turn, time_str, '-', '-', '-', '-']
                
            for col_idx, val in enumerate(row_values, 1):
                ws.cell(row=4 + turn, column=col_idx, value=val)
                
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# App Header
st.html("""
<div class="main-header">
    <h1>🏇 Umamusume Completionist Planner</h1>
    <p>Career optimization engine and schedule generator for easier "[Your Oshi] Completionist" title acquiring</p>
</div>
""")

# Fetch all races (cached in session state)
if 'all_races' not in st.session_state:
    try:
        st.session_state.all_races = get_all_races()
    except Exception as e:
        st.error(f"Error loading database: {e}")
        st.stop()
all_races = st.session_state.all_races

# Initialize persistent selection set from browser cookies or defaults
if 'selected_ids' not in st.session_state:
    cookie_val = st.context.cookies.get("uma_selected_races")
    if cookie_val is not None:
        if cookie_val.strip() == "":
            st.session_state.selected_ids = set()
        else:
            try:
                st.session_state.selected_ids = {int(x) for x in cookie_val.split(",") if x.strip()}
            except Exception:
                st.session_state.selected_ids = {r['id'] for r in all_races if r['grade'] == 0}
    else:
        st.session_state.selected_ids = {r['id'] for r in all_races if r['grade'] == 0}
        
    # Pre-populate all checkbox keys in session state once
    for r in all_races:
        st.session_state[f"cb_race_{r['id']}"] = (r['id'] in st.session_state.selected_ids)

def set_race_selection(filter_func, state):
    for r in all_races:
        if filter_func(r):
            if state:
                st.session_state.selected_ids.add(r['id'])
            else:
                st.session_state.selected_ids.discard(r['id'])
            key = f"cb_race_{r['id']}"
            if key in st.session_state:
                st.session_state[key] = state

def toggle_race(race_id):
    key = f"cb_race_{race_id}"
    if st.session_state.get(key, False):
        st.session_state.selected_ids.add(race_id)
    else:
        st.session_state.selected_ids.discard(race_id)

def save_selection_to_cookie():
    # Serialize selected IDs as a comma-separated list
    ids_str = ",".join(str(x) for x in st.session_state.selected_ids)
    # Inject a hidden iframe to write the cookie to the client browser document
    st.iframe(
        f"""
        <script>
        parent.document.cookie = "uma_selected_races={ids_str}; path=/; max-age=31536000; SameSite=Lax";
        </script>
        """,
        height=1,
        width=1
    )

# Sidebar: Inputs
st.sidebar.markdown("## ⚙️ Settings")

# Default values set to Curren Chan example (Turf A, Dirt G, Sprint A, Mile F, Medium G, Long G)
st.sidebar.markdown("### Starting Base Ranks")
base_turf = st.sidebar.selectbox("Turf", list(RANK_SCALE.keys()), index=list(RANK_SCALE.keys()).index('A'))
base_dirt = st.sidebar.selectbox("Dirt", list(RANK_SCALE.keys()), index=list(RANK_SCALE.keys()).index('G'))
base_sprint = st.sidebar.selectbox("Sprint", list(RANK_SCALE.keys()), index=list(RANK_SCALE.keys()).index('A'))
base_mile = st.sidebar.selectbox("Mile", list(RANK_SCALE.keys()), index=list(RANK_SCALE.keys()).index('F'))
base_medium = st.sidebar.selectbox("Medium", list(RANK_SCALE.keys()), index=list(RANK_SCALE.keys()).index('G'))
base_long = st.sidebar.selectbox("Long", list(RANK_SCALE.keys()), index=list(RANK_SCALE.keys()).index('G'))

base_ranks = {
    'turf': RANK_SCALE[base_turf],
    'dirt': RANK_SCALE[base_dirt],
    'sprint': RANK_SCALE[base_sprint],
    'mile': RANK_SCALE[base_mile],
    'medium': RANK_SCALE[base_medium],
    'long': RANK_SCALE[base_long]
}

st.sidebar.markdown("---")

# Consecutive races limit
st.sidebar.markdown("### Consecutive Run Limits")
limit_type = st.sidebar.radio(
    "Select Consecutive Race Rule",
    ["No Limit", "Max 2 in a row (avoid 3)", "Max 3 in a row (avoid 4)", "Custom limit"],
    index=1 # Max 2 in a row is the typical game practice
)

if limit_type == "No Limit":
    max_consecutive = None
elif limit_type == "Max 2 in a row (avoid 3)":
    max_consecutive = 2
elif limit_type == "Max 3 in a row (avoid 4)":
    max_consecutive = 3
else:
    max_consecutive = st.sidebar.number_input("Max consecutive races", min_value=1, max_value=10, value=2)

st.sidebar.markdown("---")

# Sidebar footer for AGPL v3 source code compliance
st.sidebar.markdown(
    """
    <div style="font-size: 0.85rem; color: #888; text-align: center; line-height: 1.4;">
        <strong>UmaCP v1.0.0</strong><br>
        Licensed under <a href="https://www.gnu.org/licenses/agpl-3.0.html" target="_blank" style="color: #888; text-decoration: underline;">GNU AGPL v3</a>.<br><br>
        <a href="https://github.com/1njure/uma-cp" target="_blank" style="color: #a855f7; text-decoration: none; font-weight: 600; background: rgba(168, 85, 247, 0.1); padding: 6px 12px; border-radius: 6px; border: 1px solid rgba(168, 85, 247, 0.2); display: inline-block;">
            📥 View & Download Source
        </a>
    </div>
    """,
    unsafe_allow_html=True
)

# Main Content Layout
# Tabs to separate selecting races and results
tab_select, tab_results = st.tabs(["🎯 Select Races to Complete", "📊 Generated Career Schedule"])

with tab_select:
    st.subheader("Select Outstanding Uncompleted Races")
    st.write("Toggle the checkboxes below to select the list of races you want to schedule and complete. By default, G1 races are pre-selected.")
    
    # Group races by Grade
    races_g1 = [r for r in all_races if r['grade'] == 0]
    races_g2 = [r for r in all_races if r['grade'] == 1]
    races_g3 = [r for r in all_races if r['grade'] == 2]
    
    # Selection control helpers (Buttons block)
    st.markdown("##### ➕ Quick Add to Selection")
    
    col_g1, col_g2, col_g3, col_all, col_clear = st.columns(5)
    with col_g1:
        if st.button("Add G1", width="stretch"):
            set_race_selection(lambda r: r['grade'] == 0, True)
            st.rerun()
    with col_g2:
        if st.button("Add G2", width="stretch"):
            set_race_selection(lambda r: r['grade'] == 1, True)
            st.rerun()
    with col_g3:
        if st.button("Add G3", width="stretch"):
            set_race_selection(lambda r: r['grade'] == 2, True)
            st.rerun()
    with col_all:
        if st.button("Select All", type="primary", width="stretch"):
            set_race_selection(lambda r: True, True)
            st.rerun()
    with col_clear:
        if st.button("Clear All", width="stretch"):
            set_race_selection(lambda r: True, False)
            st.rerun()
            
    col_ts, col_tm, col_tmed, col_tl = st.columns(4)
    with col_ts:
        if st.button("Add Turf Sprint", width="stretch"):
            set_race_selection(lambda r: r['surface'] == 0 and r['distance'] == 0, True)
            st.rerun()
    with col_tm:
        if st.button("Add Turf Mile", width="stretch"):
            set_race_selection(lambda r: r['surface'] == 0 and r['distance'] == 1, True)
            st.rerun()
    with col_tmed:
        if st.button("Add Turf Medium", width="stretch"):
            set_race_selection(lambda r: r['surface'] == 0 and r['distance'] == 2, True)
            st.rerun()
    with col_tl:
        if st.button("Add Turf Long", width="stretch"):
            set_race_selection(lambda r: r['surface'] == 0 and r['distance'] == 3, True)
            st.rerun()
            
    col_ds, col_dm, col_dmed, col_empty = st.columns(4)
    with col_ds:
        if st.button("Add Dirt Sprint", width="stretch"):
            set_race_selection(lambda r: r['surface'] == 1 and r['distance'] == 0, True)
            st.rerun()
    with col_dm:
        if st.button("Add Dirt Mile", width="stretch"):
            set_race_selection(lambda r: r['surface'] == 1 and r['distance'] == 1, True)
            st.rerun()
    with col_dmed:
        if st.button("Add Dirt Medium", width="stretch"):
            set_race_selection(lambda r: r['surface'] == 1 and r['distance'] == 2, True)
            st.rerun()
    with col_empty:
        st.write("")
        
    # Search filter
    search_query = st.text_input(
        "🔍 Search races",
        value=st.session_state.get('race_search_query', ''),
        placeholder="Search by name (e.g. Arima, Derby, Kawasaki, Turf)...",
        key="race_search_query",
        label_visibility="collapsed"
    )
    
    # Helper to render a group of races with a header in columns
    def render_race_group(group_title, races_list, query=""):
        # Filter races by search query
        if query:
            q = query.lower()
            filtered = [r for r in races_list if q in f"{r['name']} {['Turf','Dirt'][r['surface']]} {['Sprint','Mile','Medium','Long'][r['distance']]}".lower()]
        else:
            filtered = races_list
        
        if not filtered:
            return  # Don't show empty sections
        
        # Count selected races in this group
        selected_in_group = sum(1 for r in races_list if r['id'] in st.session_state.selected_ids)
        
        if query:
            st.markdown(f"#### {group_title} ({selected_in_group} selected, {len(filtered)} matching of {len(races_list)})")
        else:
            st.markdown(f"#### {group_title} ({selected_in_group} selected of {len(races_list)})")
        cols = st.columns(3)
        for i, race in enumerate(filtered):
            col = cols[i % 3]
            surf_str = "Turf" if race['surface'] == 0 else "Dirt"
            dist_str = ["Sprint", "Mile", "Medium", "Long"][race['distance']]
            label = f"{race['name']} ({surf_str} {dist_str})"
            
            cb_key = f"cb_race_{race['id']}"
            col.checkbox(
                label,
                key=cb_key,
                on_change=toggle_race,
                args=(race['id'],)
            )
            
    # Show how many are selected in total
    selected_count = len(st.session_state.selected_ids)
    st.info(f"Currently selected: **{selected_count}** of **{len(all_races)}** races.")
    
    st.markdown("---")
    render_race_group("🏆 Grade 1 Races (Major Milestones)", races_g1, search_query)
    st.markdown("---")
    render_race_group("🐎 Grade 2 Races (Stepping Stones)", races_g2, search_query)
    st.markdown("---")
    render_race_group("🏁 Grade 3 Races (Local Events)", races_g3, search_query)
    st.markdown("---")
    
    # Show warning if search yields no results
    if search_query:
        q = search_query.lower()
        total_matches = sum(
            1 for r in all_races
            if q in f"{r['name']} {['Turf','Dirt'][r['surface']]} {['Sprint','Mile','Medium','Long'][r['distance']]}".lower()
        )
        if total_matches == 0:
            st.warning(f"No races found matching **\"{search_query}\"**. Try a different search term.")
    
    selected_races = [r for r in all_races if r['id'] in st.session_state.selected_ids]
    
    if st.button("🚀 Calculate Optimal Schedule", type="primary", width="stretch"):
        if not selected_races:
            st.warning("Please select at least one race to schedule.")
        else:
            with st.spinner("Optimizing careers and avoiding turn collisions..."):
                # Run optimization
                careers = optimize_schedule(base_ranks, selected_races, max_consecutive)
                st.session_state.careers = careers
                st.session_state.ran_optimization = True
                
                # Automatically redirect to results tab
                st.success(f"Successfully planned all {len(selected_races)} races across {len(careers)} careers! Go to the 'Generated Career Schedule' tab above to see details.")

with tab_results:
    if 'ran_optimization' not in st.session_state or not st.session_state.ran_optimization:
        st.info("Please select your races and click 'Calculate Optimal Schedule' on the first tab to view the results.")
    else:
        careers = st.session_state.careers
        
        # Summary metrics
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            st.markdown(f"### Total Careers Required: <span class='summary-metric'>{len(careers)}</span>", unsafe_allow_html=True)
            st.write("A career represents a single run of 3 years (72 turns) in the scenario.")
        with col_m2:
            # Excel export button
            excel_data = generate_excel_file(careers, base_ranks)
            st.download_button(
                label="📥 Export Schedule to Excel (.xlsx)",
                data=excel_data,
                file_name="umamusume_race_schedule.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch"
            )
            
        st.markdown("---")
        
        # Loop through each career and display
        for idx, career in enumerate(careers):
            build = career['build']
            build_reqs = get_build_requirements(build)
            
            # Format build requirement string
            build_html = ""
            if build_reqs['total_stars'] == 0:
                build_html = "No factor inheritance upgrades required"
            else:
                badges = []
                for attr, info in build_reqs['attributes'].items():
                    base_val = base_ranks.get(attr, 0)
                    final_val = min(6, base_val + info['steps'])
                    final_letter = RANK_NAMES.get(final_val, 'G')
                    badges.append(f"<span class='stat-badge'>+{info['steps']} {attr.capitalize()} ({info['stars']}★, {info['slots']} slots, [{final_letter}])</span>")
                build_html = " &nbsp; ".join(badges)
                
            # Use Streamlit's native container with border
            with st.container(border=True):
                # Header row inside container
                col_c1, col_c2 = st.columns([1, 2])
                with col_c1:
                    st.subheader(f"🏃 Career {idx + 1}")
                with col_c2:
                    st.markdown(f"<div style='text-align: right; margin-top: 5px; font-size: 1rem;'>Build Requirements: {build_html}</div>", unsafe_allow_html=True)
                
                # Visual Calendar
                st.markdown("##### 📅 Calendar Schedule Grid")
                tab_yr1, tab_yr2, tab_yr3 = st.tabs(["🌅 Junior (Year 1)", "🐎 Classic (Year 2)", "🏆 Senior (Year 3)"])
                with tab_yr1:
                    st.html(render_calendar_year(career['schedule'], 0))
                with tab_yr2:
                    st.html(render_calendar_year(career['schedule'], 1))
                with tab_yr3:
                    st.html(render_calendar_year(career['schedule'], 2))
                
                # List of races in this career
                st.markdown("##### 🏁 Scheduled Races List")
                race_list_data = []
                for turn in range(1, 73):
                    race = career['schedule'][turn]
                    if race:
                        grade_str = ["G1", "G2", "G3"][race['grade']]
                        dist_str = ["Sprint", "Mile", "Medium", "Long"][race['distance']]
                        surf_str = ["Turf", "Dirt"][race['surface']]
                        
                        year = "Junior" if turn <= 24 else ("Classic" if turn <= 48 else "Senior")
                        turn_in_year = (turn - 1) % 24
                        month = (turn_in_year // 2) + 1
                        half = "Early" if (turn_in_year % 2) == 0 else "Late"
                        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                        
                        race_list_data.append({
                            "Turn": turn,
                            "Calendar": f"{year} {month_names[month-1]} {half}",
                            "Race Name": race['name'],
                            "Grade": grade_str,
                            "Distance": dist_str,
                            "Surface": surf_str
                        })
                        
                if race_list_data:
                    # Render schedule table
                    md_table = "| Turn | Calendar | Race Name | Grade | Distance | Surface |\n"
                    md_table += "| :--- | :--- | :--- | :--- | :--- | :--- |\n"
                    for row in race_list_data:
                        md_table += f"| {row['Turn']} | {row['Calendar']} | **{row['Race Name']}** | {row['Grade']} | {row['Distance']} | {row['Surface']} |\n"
                    st.markdown(md_table, unsafe_allow_html=True)
                else:
                    st.write("No races scheduled in this career.")

# Save selection to cookie
save_selection_to_cookie()
